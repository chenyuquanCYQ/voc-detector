"""
氣味檢測器市調資料分析系統
使用 Ollama (gemma4:e4b) 進行本地端 LLM 分析

v2 更新：
- 輸入來源改為直接從 Google Sheets 下載最新資料
- 輸出路徑改為絕對路徑，不受執行目錄影響
- 增量分析：只分析新增筆數，跳過已有 analyzed_at 的列
"""

import pandas as pd
import json
import time
import re
import sys
import io
import urllib.request
from pathlib import Path
from datetime import datetime

try:
    import ollama
    OLLAMA_AVAILABLE = True
except ImportError:
    OLLAMA_AVAILABLE = False
    print("⚠️  ollama 套件未安裝，請執行: pip install ollama")

# ═══════════════════════════════════════════════════════════
#  ★ 設定區 - 只需修改這裡
# ═══════════════════════════════════════════════════════════

# Google 試算表 ID（網址 /d/ 後面那串，試算表需設為「知道連結的人可檢視」）
SHEET_ID = "1AyPfZdyqrkIwsLaKh7BNLqlgRhcXOXT9vrakUoeos0I"

# 分析結果輸出路徑（絕對路徑，不受執行目錄影響）
OUTPUT_FILE = r"D:\02-AIProject\VOCsDetector\氣味檢測器市調_分析結果.xlsx"

# 模型設定
MODEL_NAME   = "gemma4:e4b"
BATCH_DELAY  = 0.5    # 每筆請求間隔（秒）
MAX_RETRIES  = 2      # LLM 失敗重試次數
FORCE_RERUN  = False  # True = 強制重新分析所有列

# ═══════════════════════════════════════════════════════════

# ── Prompt 設計 ──────────────────────────────────────────────────
SYSTEM_PROMPT = """你是一位精通氣味檢測器與電子鼻產業的市場調查分析師。
請根據產品描述，嚴格以 JSON 格式輸出分析結果，不得加入任何前言或說明文字。"""

def build_extraction_prompt(product_name: str, brand: str, description: str) -> str:
    return f"""請分析以下氣味檢測器產品，並以 JSON 格式輸出，欄位定義如下：

產品名稱：{product_name}
品牌/公司：{brand}
產品描述：{description}

輸出 JSON 格式（所有值請用繁體中文或英文縮寫，不可為 null）：
{{
  "sensor_type": "感測器技術類型，如 MOS/MEMS/GC-MS/NDIR/電化學/光離子/陣列式/混合型/不明",
  "form_factor": "產品型態，如 手持可攜式/固定式/嵌入式模組/無人機搭載/桌面設備/消費穿戴/不明",
  "precision_tier": "精度等級：實驗室級/工業級/消費電子級/不明",
  "trl": "技術成熟度：原型/研發中/小量商用/成熟商用",
  "target_gases": ["目標氣體清單，如 VOCs/H2S/NH3/丙酮/乙醇/CO/NO2/甲醛/多種/不明"],
  "output_type": "數據輸出：分級顯示/精確數值(ppm-ppb)/氣味指紋圖譜/多種輸出/不明",
  "ecosystem": "生態系整合：無/藍牙App/IoT雲端/AI驅動/SaaS平台/多種整合",
  "application_segments": ["應用場景清單，如 食品品質/環境監測/工業安全/醫療健康/智慧家居/農業/國防/不明"],
  "competitive_moat": "競爭護城河：純硬體/軟硬整合/資料平台/演算法IP/標準化認證/不明",
  "key_features": "最多3個獨特功能關鍵詞，以逗號分隔",
  "confidence": "分析信心度：高/中/低（依描述詳細程度而定）"
}}"""


# ── Google Sheets 下載 ───────────────────────────────────────────
def fetch_gsheet(sheet_id: str) -> pd.DataFrame:
    """直接從 Google Sheets 下載最新資料（不需 API Key）"""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid=0"
    print(f"📡 從 Google Sheets 下載最新資料...")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            df = pd.read_csv(io.BytesIO(resp.read()))
        print(f"  ✅ 雲端取得 {len(df)} 筆")
        return df
    except Exception as e:
        print(f"  ❌ Google Sheets 下載失敗：{e}")
        print(f"     請確認試算表已設為「知道連結的人可以檢視」")
        return pd.DataFrame()


# ── 讀取已有分析結果（增量用）───────────────────────────────────
def load_existing_results(output_path: str) -> pd.DataFrame:
    """讀取已有的分析結果，用於增量比對"""
    p = Path(output_path)
    if not p.exists():
        print(f"  ℹ️  尚無分析結果檔案，將從頭開始分析")
        return pd.DataFrame()
    try:
        df = pd.read_excel(output_path, sheet_name="原始+分析")
        print(f"  ✅ 讀入既有分析結果 {len(df)} 筆")
        return df
    except Exception as e:
        print(f"  ⚠️  讀取既有結果失敗：{e}")
        return pd.DataFrame()


# ── 合併雲端新資料與既有分析結果 ────────────────────────────────
def merge_with_existing(df_cloud: pd.DataFrame,
                        df_existing: pd.DataFrame) -> pd.DataFrame:
    """
    以產品名稱為 key：
    - 既有分析結果保留（已有 analyzed_at）
    - 雲端有但既有沒有的，補入為待分析列
    """
    ANALYSIS_COLS_EMPTY = {col: "" for col in ANALYSIS_COLS}

    if df_existing.empty:
        # 全新開始，直接用雲端資料，補入分析欄位
        for col in ANALYSIS_COLS:
            if col not in df_cloud.columns:
                df_cloud[col] = ""
        return df_cloud

    # 找欄位名稱
    name_col_cloud = next(
        (c for c in df_cloud.columns if "產品" in c and "名" in c), df_cloud.columns[0]
    )
    name_col_exist = next(
        (c for c in df_existing.columns if "產品" in c and "名" in c), df_existing.columns[0]
    )

    existing_names = set(
        str(v).lower().strip()
        for v in df_existing[name_col_exist]
        if pd.notna(v)
    )

    # 找出雲端新增但既有沒有的列
    new_rows = df_cloud[
        ~df_cloud[name_col_cloud].apply(
            lambda v: str(v).lower().strip() in existing_names
        )
    ].copy()

    if len(new_rows) > 0:
        for col in ANALYSIS_COLS:
            if col not in new_rows.columns:
                new_rows[col] = ""
        print(f"  ☁️  發現 {len(new_rows)} 筆新增資料，加入待分析佇列")
    else:
        print(f"  ✅ 無新增資料")

    # 確保 existing 有所有 analysis 欄位
    for col in ANALYSIS_COLS:
        if col not in df_existing.columns:
            df_existing[col] = ""

    # 合併
    combined = pd.concat([df_existing, new_rows], ignore_index=True)
    return combined


# ── LLM 呼叫 ────────────────────────────────────────────────────
def call_llm(prompt: str, retries: int = MAX_RETRIES) -> dict | None:
    if not OLLAMA_AVAILABLE:
        return None
    for attempt in range(retries + 1):
        try:
            response = ollama.chat(
                model=MODEL_NAME,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": prompt},
                ],
                options={"temperature": 0.1, "num_predict": 2048},
                think=False,
            )
            raw = response["message"]["content"].strip()
            if not raw:
                print(f"   ⚠️  模型回傳空白（嘗試 {attempt+1}）")
                time.sleep(2)
                continue
            json_match = re.search(r"\{.*\}", raw, re.DOTALL)
            if json_match:
                return json.loads(json_match.group())
            print(f"   ⚠️  找不到 JSON（嘗試 {attempt+1}）：{raw[:80]}")
        except json.JSONDecodeError as e:
            print(f"   ⚠️  JSON 解析失敗（嘗試 {attempt+1}）：{e}")
        except Exception as e:
            print(f"   ⚠️  LLM 錯誤（嘗試 {attempt+1}）：{e}")
            time.sleep(2)
    return None

# ── 分析結果欄位 ─────────────────────────────────────────────────
ANALYSIS_COLS = [
    "sensor_type", "form_factor", "precision_tier", "trl",
    "target_gases", "output_type", "ecosystem",
    "application_segments", "competitive_moat",
    "key_features", "confidence", "analyzed_at",
]

def flatten_list_fields(result: dict) -> dict:
    for key in list(result.keys()):
        val = result[key]
        if isinstance(val, list):
            result[key] = ", ".join(str(v) for v in val)
        elif isinstance(val, dict):
            result[key] = json.dumps(val, ensure_ascii=False)
        elif val is None:
            result[key] = ""
        else:
            result[key] = str(val)
    return result


# ── 主流程 ───────────────────────────────────────────────────────
def main():
    print(f"\n{'='*52}")
    print(f"  analyze_market.py v2  ({MODEL_NAME})")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*52}")

    # 1. 從 Google Sheets 下載最新原始資料
    df_cloud = fetch_gsheet(SHEET_ID)
    if df_cloud.empty:
        sys.exit("❌ 無法取得雲端資料，中止")

    # 2. 讀取既有分析結果（用於增量）
    df_existing = load_existing_results(OUTPUT_FILE)

    # 3. 合併：保留既有分析 + 補入新增資料
    df = merge_with_existing(df_cloud, df_existing)
    print(f"\n📊 合併後共 {len(df)} 筆，開始 LLM 分析...")

    # 4. 找欄位對應
    col_map = {
        "product_name": next((c for c in df.columns if "產品" in c and "名" in c), df.columns[0]),
        "brand":        next((c for c in df.columns if "品牌" in c or "公司" in c), df.columns[1] if len(df.columns) > 1 else df.columns[0]),
        "description":  next((c for c in df.columns if "描述" in c or "特色" in c), df.columns[2] if len(df.columns) > 2 else df.columns[0]),
    }
    print(f"📌 欄位對應：{col_map}")

    # 5. 逐筆分析（跳過已分析列）
    total = len(df)
    skipped = 0
    analyzed = 0

    for idx, row in df.iterrows():
        if not FORCE_RERUN and pd.notna(row.get("analyzed_at")) and str(row.get("analyzed_at")).strip():
            skipped += 1
            continue

        product_name = str(row.get(col_map["product_name"], "")).strip()
        brand        = str(row.get(col_map["brand"], "")).strip()
        description  = str(row.get(col_map["description"], "")).strip()

        if not description or description.lower() in ("nan", ""):
            df.at[idx, "confidence"]  = "低"
            df.at[idx, "analyzed_at"] = datetime.now().isoformat(timespec="seconds")
            continue

        prompt = build_extraction_prompt(product_name, brand, description)
        result = call_llm(prompt)

        if result:
            result = flatten_list_fields(result)
            result["analyzed_at"] = datetime.now().isoformat(timespec="seconds")
            for col in ANALYSIS_COLS:
                df.at[idx, col] = result.get(col, "")
            status = f"[{result.get('sensor_type','?')}] [{result.get('form_factor','?')}]"
            analyzed += 1
        else:
            df.at[idx, "confidence"]  = "分析失敗"
            df.at[idx, "analyzed_at"] = datetime.now().isoformat(timespec="seconds")
            status = "⚠️ 分析失敗"

        print(f"  [{idx+1:>4}/{total}] {product_name[:30]:<30} → {status}")
        time.sleep(BATCH_DELAY)

    print(f"\n📊 完成！本次分析：{analyzed} 筆，跳過（已有結果）：{skipped} 筆")

    # 6. 儲存結果
    _write_output(df)
    print(f"💾 已儲存至：{OUTPUT_FILE}")


def _write_output(df: pd.DataFrame):
    Path(OUTPUT_FILE).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="原始+分析", index=False)

        summary_data = {}
        for col in ("sensor_type", "form_factor", "precision_tier", "trl",
                    "output_type", "ecosystem", "competitive_moat"):
            if col in df.columns:
                summary_data[col] = df[col].value_counts().reset_index()
                summary_data[col].columns = [col, "數量"]

        row_offset = 0
        summary_sheet = writer.book.create_sheet("統計摘要")
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", start_color="2F5496")
        center_align = Alignment(horizontal="center")

        for col_name, tbl in summary_data.items():
            summary_sheet.cell(row=row_offset+1, column=1, value=col_name).font = Font(bold=True, size=12)
            summary_sheet.cell(row=row_offset+1, column=1).fill = PatternFill("solid", start_color="D6E4F0")
            for c_idx, header in enumerate(tbl.columns, 1):
                cell = summary_sheet.cell(row=row_offset+2, column=c_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            for r_idx, data_row in tbl.iterrows():
                for c_idx, val in enumerate(data_row, 1):
                    summary_sheet.cell(row=row_offset+3+r_idx, column=c_idx, value=val)
            row_offset += len(tbl) + 5

        summary_sheet.column_dimensions["A"].width = 28
        summary_sheet.column_dimensions["B"].width = 10


if __name__ == "__main__":
    main()